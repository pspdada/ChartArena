"""ChartArena benchmark 评分结果分析脚本。

读取 judge.py 的评分结果，按多种维度汇总统计，生成 Excel 报表。

功能：
    1. 任务总览（主 Excel 第一个 Sheet）：每个模型在每个 task 的总平均分
    2. 每个 task 的总分表（主 Excel 其余 Sheet）：模型 × 各文件的评分
    3. 按图表类型分页：每个 task 一个独立 Excel，每个 chart_type 一个 Sheet
    4. 详细分类结果：detail_by_category/{model}/{task}.xlsx，按 (chart_type, img_type, lang_type) 细分

用法:
    python analyze.py

    # 保存精确 AP 指标
    python analyze.py --save_ap

    # 不按图表类型拆分
    python analyze.py --no_split_by_type

    # 不保存详细分类
    python analyze.py --no_detail
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT))

DEFAULT_INPUT_DIR = REPO_ROOT / "judge_outputs"
DEFAULT_OUTPUT_FILE = REPO_ROOT / "judge_outputs" / "results_analysis.xlsx"
DETAIL_DIR_NAME = "detail_by_category"
BY_TYPE_DIR_NAME = "by_chart_type"
MAX_WORKERS = 32

# ============================================================
# 显示格式配置
# ============================================================
DISPLAY_DECIMALS = 3
NUMBER_FORMAT = "0." + "0" * DISPLAY_DECIMALS

MAP_METRICS = ["em", "map_strict", "map_slight", "map_high"]
AP_METRICS = [
    "ap_50_strict",
    "ap_75_strict",
    "ap_90_strict",
    "ap_50_slight",
    "ap_75_slight",
    "ap_90_slight",
    "ap_50_high",
    "ap_75_high",
    "ap_90_high",
]
ALL_METRICS = MAP_METRICS + AP_METRICS

CHART_TYPE_PRIORITY = [
    "柱状图",
    "折线图",
    "饼图",
    "雷达图",
    "箱线图",
    "组合图",
    "流程图",
    "思维导图",
    "散点图",
    "面积图",
    "热力图",
    "树状图",
    "漏斗图",
]
LANG_TYPE_PRIORITY = ["英文", "中文"]
TASK_PRIORITY = [
    "SE_MD",
    "SE_JSON",
    "SE_CSV",
    "SE_CODE",
    "SE_SVG",
    "SE_MERMAID",
    "SE_GRAPHVIZ",
    "SE_PLANTUML",
    "SE_DIAGRAMS",
    "SE_D2",
    "SE_CYTOSCAPE",
]

DIM_PRIORITY_MAP: dict[str, list[str]] = {
    "chart_type": CHART_TYPE_PRIORITY,
    "lang_type": LANG_TYPE_PRIORITY,
}

IMG_TYPE_RENAME_MAP: dict[str, str] = {
    "随拍手写": "随拍手绘",
    "电子印刷": "扫描印刷",
}


def _priority_index(value: str, priority_list: list[str]) -> tuple[int, str]:
    try:
        return (priority_list.index(value), value)
    except ValueError:
        return (len(priority_list), value)


def _sort_key_by_dims(row_values: tuple, dim_keys: list[str]) -> tuple:
    parts: list = []
    for val, key in zip(row_values, dim_keys):
        priority = DIM_PRIORITY_MAP.get(key, [])
        parts.append(_priority_index(str(val), priority))
    return tuple(parts)


def parse_args():
    p = argparse.ArgumentParser(description="分析 ChartArena 评分结果并生成 Excel 报表")
    p.add_argument("--input_dir", type=str, default=str(DEFAULT_INPUT_DIR), help="评分结果输入目录")
    p.add_argument("--output", type=str, default=None, help="输出 Excel 文件路径（默认在 input_dir 下）")
    p.add_argument("--save_ap", action="store_true", default=False, help="是否保存精确的 AP 指标")
    p.add_argument("--no_split_by_type", action="store_true", default=False, help="不按图表类型分页保存")
    p.add_argument("--no_detail", action="store_true", default=False, help="不保存详细分类结果")
    return p.parse_args()


# ============================================================
# 数据解析
# ============================================================


def _parse_single_jsonl(file_path: str, model_name: str, file_stem: str) -> tuple[list[dict], str | None]:
    rows: list[dict] = []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    item = json.loads(line.strip())
                except json.JSONDecodeError:
                    continue
                item["_model_name"] = model_name
                item["_source_jsonl"] = file_stem
                old_it = item.get("img_type", "")
                if old_it in IMG_TYPE_RENAME_MAP:
                    item["img_type"] = IMG_TYPE_RENAME_MAP[old_it]
                rows.append(item)
        return rows, None
    except Exception as e:
        return rows, f"读取 {file_path} 出错: {e}"


def parse_all_judge_files(input_dir: str, max_workers: int = MAX_WORKERS) -> list[dict]:
    """解析输入目录下所有模型的评分结果（多线程并发读取）。"""
    skip_dirs = {DETAIL_DIR_NAME, BY_TYPE_DIR_NAME}
    model_dirs = sorted(
        d
        for d in os.listdir(input_dir)
        if os.path.isdir(os.path.join(input_dir, d)) and d not in skip_dirs and not d.startswith((".", "_"))
    )
    print(f"找到 {len(model_dirs)} 个模型目录")

    tasks: list[tuple[str, str, str]] = []
    for model_name in model_dirs:
        model_path = os.path.join(input_dir, model_name)
        jsonl_files = sorted([f for f in os.listdir(model_path) if f.endswith(".jsonl")])
        print(f"  模型 {model_name}: {len(jsonl_files)} 个文件")
        for jsonl_file in jsonl_files:
            file_path = os.path.join(model_path, jsonl_file)
            file_stem = os.path.splitext(jsonl_file)[0]
            tasks.append((file_path, model_name, file_stem))

    if not tasks:
        print("共读取 0 条数据\n")
        return []

    workers = min(max_workers, len(tasks))
    print(f"使用 {workers} 个线程并发读取 {len(tasks)} 个文件...")

    all_rows: list[dict] = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_task = {executor.submit(_parse_single_jsonl, fp, mn, fs): (fp, mn, fs) for fp, mn, fs in tasks}
        with tqdm(total=len(tasks), desc="读取 jsonl", ncols=100) as pbar:
            for future in as_completed(future_to_task):
                fp, mn, fs = future_to_task[future]
                try:
                    rows, err = future.result()
                except Exception as e:
                    rows, err = [], f"读取 {fp} 出错: {e}"
                if err:
                    tqdm.write(f"    {err}")
                all_rows.extend(rows)
                pbar.update(1)

    print(f"共读取 {len(all_rows)} 条数据\n")
    return all_rows


def extract_scores_per_task(item: dict) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for task_name, task_result in (item.get("judge_results") or {}).items():
        score_dict = task_result.get("score", {})
        if isinstance(score_dict, dict) and "map_strict" in score_dict:
            result[task_name] = {k: score_dict.get(k, 0.0) for k in ALL_METRICS}
    return result


def collect_all_tasks(all_rows: list[dict]) -> list[str]:
    tasks = set()
    for item in all_rows:
        tasks.update(extract_scores_per_task(item).keys())
    sorted_tasks = [t for t in TASK_PRIORITY if t in tasks]
    for t in sorted(tasks):
        if t not in sorted_tasks:
            sorted_tasks.append(t)
    return sorted_tasks


# ============================================================
# 原子桶（一次扫描，所有视图均可 rollup 得出）
# ============================================================

ATOMIC_KEYS: list[str] = ["_model_name", "_source_jsonl", "chart_type", "img_type", "lang_type"]


def _build_atomic_buckets(rows: list[dict]) -> tuple[dict, dict]:
    atomic: dict[str, dict[tuple, dict[str, list]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
    )
    row_index_to_atomic: dict[int, tuple] = {}
    for idx, item in enumerate(rows):
        scores_per_task = extract_scores_per_task(item)
        atomic_key = tuple(item.get(k, "未知") for k in ATOMIC_KEYS)
        row_index_to_atomic[idx] = atomic_key
        if not scores_per_task:
            continue
        for task_name, scores in scores_per_task.items():
            task_map = atomic[task_name][atomic_key]
            for metric, value in scores.items():
                slot = task_map[metric]
                slot[0] += float(value)
                slot[1] += 1

    frozen: dict[str, dict[tuple, dict[str, tuple[float, int]]]] = {}
    for task_name, group_map in atomic.items():
        frozen_task: dict[tuple, dict[str, tuple[float, int]]] = {}
        for gk, mm in group_map.items():
            frozen_task[gk] = {m: (s, c) for m, (s, c) in mm.items()}
        frozen[task_name] = frozen_task
    return frozen, row_index_to_atomic


def _atomic_key_to_dict(atomic_key: tuple) -> dict[str, str]:
    return {k: v for k, v in zip(ATOMIC_KEYS, atomic_key)}


def _rollup_from_atomic(atomic_for_task: dict, group_keys: list[str], atomic_key_filter=None) -> dict[tuple, dict]:
    key_indices = [ATOMIC_KEYS.index(k) for k in group_keys]
    agg: dict[tuple, dict[str, list]] = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
    for atomic_key, metric_map in atomic_for_task.items():
        if atomic_key_filter is not None and not atomic_key_filter(_atomic_key_to_dict(atomic_key)):
            continue
        group_value = tuple(atomic_key[i] for i in key_indices)
        bucket = agg[group_value]
        for metric, (s, c) in metric_map.items():
            slot = bucket[metric]
            slot[0] += s
            slot[1] += c
    result: dict[tuple, dict] = {}
    for group_value, metric_map in agg.items():
        avg: dict = {"count": 0}
        max_count = 0
        for metric in ALL_METRICS:
            sc = metric_map.get(metric)
            if sc and sc[1] > 0:
                avg[metric] = sc[0] / sc[1]
                if sc[1] > max_count:
                    max_count = sc[1]
            else:
                avg[metric] = 0.0
        avg["count"] = max_count
        result[group_value] = avg
    return result


def sort_chart_types(chart_types: list[str]) -> list[str]:
    sorted_types = [ct for ct in CHART_TYPE_PRIORITY if ct in chart_types]
    for ct in sorted(chart_types):
        if ct not in sorted_types:
            sorted_types.append(ct)
    return sorted_types


# ============================================================
# Excel 写入工具
# ============================================================


def get_metric_columns(save_ap: bool) -> list[str]:
    return ALL_METRICS if save_ap else MAP_METRICS


def style_header(worksheet, max_col: int, header_rows: int = 2, fill_color: str = "CCE5FF"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, header_rows + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment


def auto_column_width(worksheet, min_width: int = 10, max_width: int = 30):
    for col_cells in worksheet.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        width = max(min_width, min(max_len + 2, max_width))
        worksheet.column_dimensions[col_letter].width = width


def apply_number_format(worksheet, data_start_row: int, col_indices: list[int], number_format: str = NUMBER_FORMAT):
    if not col_indices:
        return
    max_row = worksheet.max_row
    for row in range(data_start_row, max_row + 1):
        for col in col_indices:
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = number_format


def safe_sheet_name(name: str) -> str:
    for ch in [":", "/", "\\", "?", "*", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:31]


# ============================================================
# 表构造
# ============================================================


def build_task_summary_sheet(atomic, task_name, save_ap, atomic_key_filter=None):
    metrics = get_metric_columns(save_ap)
    atomic_for_task = atomic.get(task_name, {})
    if not atomic_for_task:
        return pd.DataFrame(), [], []

    group_avgs = _rollup_from_atomic(atomic_for_task, ["_model_name", "_source_jsonl"], atomic_key_filter)
    model_overall = _rollup_from_atomic(atomic_for_task, ["_model_name"], atomic_key_filter)
    if not group_avgs and not model_overall:
        return pd.DataFrame(), [], []

    all_models = sorted({gk[0] for gk in model_overall.keys()})
    all_files = sorted({gk[1] for gk in group_avgs.keys()})

    rows_data = []
    for model in all_models:
        row = {"模型名称": model}
        overall = model_overall.get((model,), {})
        row["平均分_count"] = overall.get("count", 0)
        for m in metrics:
            row[f"平均分_{m}"] = overall.get(m, 0.0)
        for file_stem in all_files:
            file_avg = group_avgs.get((model, file_stem), {})
            row[f"{file_stem}_count"] = file_avg.get("count", 0)
            for m in metrics:
                row[f"{file_stem}_{m}"] = file_avg.get(m, 0.0)
        rows_data.append(row)

    df = pd.DataFrame(rows_data)
    header1 = ["模型名称"] + ["平均分"] * (len(metrics) + 1)
    for file_stem in all_files:
        header1.extend([file_stem] * (len(metrics) + 1))
    header2 = ["模型名称", "count"] + metrics
    for _ in all_files:
        header2.append("count")
        header2.extend(metrics)

    col_order = ["模型名称", "平均分_count"] + [f"平均分_{m}" for m in metrics]
    for file_stem in all_files:
        col_order.append(f"{file_stem}_count")
        col_order.extend(f"{file_stem}_{m}" for m in metrics)
    for c in col_order:
        if c not in df.columns:
            df[c] = 0.0
    df = df[col_order]
    return df, header1, header2


def write_summary_sheet(writer, df, header1, header2, sheet_name):
    if df.empty:
        return
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    ws = writer.sheets[sheet_name]
    for col_idx, value in enumerate(header1, start=1):
        ws.cell(row=1, column=col_idx, value=value)
    for col_idx, value in enumerate(header2, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    col_idx = 1
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    col_idx += 1
    i = 1
    while i < len(header1):
        group_name = header1[i]
        span = 0
        j = i
        while j < len(header1) and header1[j] == group_name:
            span += 1
            j += 1
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + span - 1)
        col_idx += span
        i = j
    style_header(ws, len(header2))
    auto_column_width(ws)
    metric_col_idxs = [idx for idx, name in enumerate(header2, start=1) if name not in ("模型名称", "count")]
    apply_number_format(ws, data_start_row=3, col_indices=metric_col_idxs)


def build_task_overview_sheet(atomic, tasks, save_ap):
    metrics = get_metric_columns(save_ap)
    per_task_overall: dict[str, dict] = {}
    all_models_set: set[str] = set()
    for task_name, atomic_for_task in atomic.items():
        rolled = _rollup_from_atomic(atomic_for_task, ["_model_name"])
        per_task_overall[task_name] = rolled
        all_models_set.update(gk[0] for gk in rolled.keys())

    all_models = sorted(all_models_set)
    rows_data = []
    for model in all_models:
        row = {"模型名称": model}
        for task_name in tasks:
            task_overall = per_task_overall.get(task_name, {})
            overall = task_overall.get((model,), {})
            row[f"{task_name}_count"] = overall.get("count", 0)
            for m in metrics:
                row[f"{task_name}_{m}"] = overall.get(m, 0.0)
        rows_data.append(row)

    df = pd.DataFrame(rows_data)
    header1 = ["模型名称"]
    for task_name in tasks:
        header1.extend([task_name] * (len(metrics) + 1))
    header2 = ["模型名称"]
    for _ in tasks:
        header2.append("count")
        header2.extend(metrics)

    col_order = ["模型名称"]
    for task_name in tasks:
        col_order.append(f"{task_name}_count")
        col_order.extend(f"{task_name}_{m}" for m in metrics)
    for c in col_order:
        if c not in df.columns:
            df[c] = 0.0
    df = df[col_order]
    return df, header1, header2


# ============================================================
# 详细分类结果
# ============================================================


def build_detail_for_model_task(atomic, model_name, task_name, save_ap):
    metrics = get_metric_columns(save_ap)
    dim_keys = ["chart_type", "img_type", "lang_type"]
    atomic_for_task = atomic.get(task_name, {})
    group_avgs = _rollup_from_atomic(atomic_for_task, dim_keys, lambda d: d.get("_model_name") == model_name)
    rows_data = []
    for (chart_type, img_type, lang_type), avg in sorted(
        group_avgs.items(), key=lambda kv: _sort_key_by_dims(kv[0], dim_keys)
    ):
        row = {"chart_type": chart_type, "img_type": img_type, "lang_type": lang_type, "count": avg.get("count", 0)}
        for m in metrics:
            row[m] = avg.get(m, 0.0)
        rows_data.append(row)
    if not rows_data:
        return pd.DataFrame()
    df = pd.DataFrame(rows_data)
    return df[["chart_type", "img_type", "lang_type", "count"] + metrics]


def _build_single_key_summary(atomic, model_name, key, task_name, metrics):
    atomic_for_task = atomic.get(task_name, {})
    group_avgs = _rollup_from_atomic(atomic_for_task, [key], lambda d: d.get("_model_name") == model_name)
    if not group_avgs:
        return pd.DataFrame()
    rows_data = []
    for (val,), avg in sorted(group_avgs.items(), key=lambda kv: _sort_key_by_dims(kv[0], [key])):
        row = {key: val, "count": avg.get("count", 0)}
        for m in metrics:
            row[m] = avg.get(m, 0.0)
        rows_data.append(row)
    df = pd.DataFrame(rows_data)
    return df[[key, "count"] + metrics]


def _build_two_key_summary(atomic, model_name, key1, key2, task_name, metrics):
    atomic_for_task = atomic.get(task_name, {})
    group_avgs = _rollup_from_atomic(atomic_for_task, [key1, key2], lambda d: d.get("_model_name") == model_name)
    if not group_avgs:
        return pd.DataFrame()
    rows_data = []
    for (v1, v2), avg in sorted(group_avgs.items(), key=lambda kv: _sort_key_by_dims(kv[0], [key1, key2])):
        row = {key1: v1, key2: v2, "count": avg.get("count", 0)}
        for m in metrics:
            row[m] = avg.get(m, 0.0)
        rows_data.append(row)
    df = pd.DataFrame(rows_data)
    return df[[key1, key2, "count"] + metrics]


def _write_sheet_with_transpose(writer, df, sheet_name, fill_color, transpose_fill_color="FFF2CC"):
    if df.empty:
        return
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    style_header(ws, len(df.columns), header_rows=1, fill_color=fill_color)
    auto_column_width(ws)
    metric_col_idxs = [idx for idx, col_name in enumerate(df.columns, start=1) if col_name in ALL_METRICS]
    apply_number_format(ws, data_start_row=2, col_indices=metric_col_idxs)

    # 转置表
    dim_cols = [c for c in df.columns if c not in ("count",) and c not in ALL_METRICS]
    if not dim_cols:
        dim_cols = [df.columns[0]]
    row_labels = [" | ".join(str(df.iloc[i][c]) for c in dim_cols) for i in range(len(df))]
    transposed = {"字段": list(df.columns)}
    for i in range(len(df)):
        transposed[row_labels[i]] = [df.iloc[i][c] for c in df.columns]
    df_t = pd.DataFrame(transposed)
    t_sheet_name = safe_sheet_name(f"{sheet_name}(T)")
    df_t.to_excel(writer, sheet_name=t_sheet_name, index=False)
    ws_t = writer.sheets[t_sheet_name]
    style_header(ws_t, len(df_t.columns), header_rows=1, fill_color=transpose_fill_color)
    auto_column_width(ws_t)
    data_col_idxs = list(range(2, len(df_t.columns) + 1))
    for r_offset, field_name in enumerate(df_t["字段"].tolist()):
        if field_name in ALL_METRICS:
            row_num = 2 + r_offset
            for col in data_col_idxs:
                cell = ws_t.cell(row=row_num, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = NUMBER_FORMAT


def _write_one_detail_xlsx(model_name, task_name, atomic, output_file, save_ap):
    try:
        metrics = get_metric_columns(save_ap)
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            _write_sheet_with_transpose(
                writer, build_detail_for_model_task(atomic, model_name, task_name, save_ap), "完整分类", "CCE5FF"
            )
            _write_sheet_with_transpose(
                writer,
                _build_single_key_summary(atomic, model_name, "chart_type", task_name, metrics),
                "按chart_type",
                "D5F5E3",
            )
            _write_sheet_with_transpose(
                writer,
                _build_single_key_summary(atomic, model_name, "img_type", task_name, metrics),
                "按img_type",
                "FCE4D6",
            )
            _write_sheet_with_transpose(
                writer,
                _build_single_key_summary(atomic, model_name, "lang_type", task_name, metrics),
                "按lang_type",
                "E2EFDA",
            )
            _write_sheet_with_transpose(
                writer,
                _build_two_key_summary(atomic, model_name, "chart_type", "img_type", task_name, metrics),
                "chart_type×img_type",
                "D9E2F3",
            )
            _write_sheet_with_transpose(
                writer,
                _build_two_key_summary(atomic, model_name, "chart_type", "lang_type", task_name, metrics),
                "chart_type×lang_type",
                "E8DAEF",
            )
        return model_name, task_name, output_file, None
    except Exception as e:
        return model_name, task_name, output_file, f"{type(e).__name__}: {e}"


def write_detail_excel(atomic, output_dir, tasks, save_ap, max_workers=MAX_WORKERS):
    os.makedirs(output_dir, exist_ok=True)
    task_to_models: dict[str, set[str]] = {}
    all_models_set: set[str] = set()
    for task_name in tasks:
        atomic_for_task = atomic.get(task_name)
        if not atomic_for_task:
            continue
        models_in_task = {gk[ATOMIC_KEYS.index("_model_name")] for gk in atomic_for_task.keys()}
        task_to_models[task_name] = models_in_task
        all_models_set.update(models_in_task)

    jobs: list[tuple[str, str, str]] = []
    for model_name in sorted(all_models_set):
        model_out_dir = os.path.join(output_dir, model_name)
        os.makedirs(model_out_dir, exist_ok=True)
        for task_name in tasks:
            if model_name not in task_to_models.get(task_name, set()):
                continue
            jobs.append((model_name, task_name, os.path.join(model_out_dir, f"{task_name}.xlsx")))

    if not jobs:
        print("  (没有需要生成的详细分类文件)")
        return

    workers = min(max_workers, len(jobs))
    print(f"  使用 {workers} 个线程并发生成 {len(jobs)} 个详细分类文件...")
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_job = {
            executor.submit(_write_one_detail_xlsx, mn, tn, atomic, out, save_ap): (mn, tn, out) for mn, tn, out in jobs
        }
        results: dict[tuple[str, str], tuple[str, str | None]] = {}
        with tqdm(total=len(jobs), desc="详细分类", ncols=100) as pbar:
            for future in as_completed(future_to_job):
                mn, tn, out = future_to_job[future]
                try:
                    _, _, _, err = future.result()
                except Exception as e:
                    err = f"{type(e).__name__}: {e}"
                results[(mn, tn)] = (out, err)
                pbar.update(1)
        for mn, tn, _out in jobs:
            out, err = results[(mn, tn)]
            if err:
                print(f"    ✗ {mn}/{tn}: {err}")
            else:
                print(f"    ✓ {mn}/{tn}: {out}")


# ============================================================
# 按图表类型分页
# ============================================================


def _write_one_by_chart_type_xlsx(task_name, atomic, sorted_types, output_file, save_ap):
    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df, h1, h2 = build_task_summary_sheet(atomic, task_name, save_ap)
            if not df.empty:
                write_summary_sheet(writer, df, h1, h2, sheet_name="总分")
            for chart_type in sorted_types:
                df_ct, ct_h1, ct_h2 = build_task_summary_sheet(
                    atomic,
                    task_name,
                    save_ap,
                    atomic_key_filter=lambda d, ct=chart_type: d.get("chart_type") == ct,
                )
                if df_ct.empty:
                    continue
                write_summary_sheet(writer, df_ct, ct_h1, ct_h2, sheet_name=safe_sheet_name(chart_type))
        return task_name, output_file, None
    except Exception as e:
        return task_name, output_file, f"{type(e).__name__}: {e}"


def write_by_chart_type_excel(atomic, output_dir, tasks, save_ap, max_workers=MAX_WORKERS):
    os.makedirs(output_dir, exist_ok=True)
    ct_idx = ATOMIC_KEYS.index("chart_type")
    chart_types_set: set[str] = set()
    for atomic_for_task in atomic.values():
        for gk in atomic_for_task.keys():
            chart_types_set.add(gk[ct_idx])
    sorted_types = sort_chart_types(sorted(chart_types_set))

    jobs = [(tn, os.path.join(output_dir, f"{tn}.xlsx")) for tn in tasks if atomic.get(tn)]
    if not jobs:
        print("  (没有需要生成的按图表类型拆分文件)")
        return

    workers = min(max_workers, len(jobs))
    print(f"  使用 {workers} 个线程并发生成 {len(jobs)} 个按图表类型拆分文件...")
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_job = {
            executor.submit(_write_one_by_chart_type_xlsx, tn, atomic, sorted_types, out, save_ap): (tn, out)
            for tn, out in jobs
        }
        results: dict[str, tuple[str, str | None]] = {}
        with tqdm(total=len(jobs), desc="按图表类型", ncols=100) as pbar:
            for future in as_completed(future_to_job):
                tn, out = future_to_job[future]
                try:
                    _, _, err = future.result()
                except Exception as e:
                    err = f"{type(e).__name__}: {e}"
                results[tn] = (out, err)
                pbar.update(1)
        for tn, _out in jobs:
            out, err = results[tn]
            if err:
                print(f"    ✗ {tn}: {err}")
            else:
                print(f"    ✓ {tn}: {out}")


# ============================================================
# 主函数
# ============================================================


def main():
    args = parse_args()

    print("=" * 72)
    print("ChartArena 评分结果分析")
    print("=" * 72)
    print(f"输入目录: {args.input_dir}")
    print(f"保存精确 AP: {args.save_ap}")
    print(f"按图表类型分页: {not args.no_split_by_type}")
    print(f"保存详细分类: {not args.no_detail}")
    print()

    if not os.path.isdir(args.input_dir):
        raise SystemExit(f"输入目录不存在: {args.input_dir}")

    output_file = args.output or os.path.join(args.input_dir, "results_analysis.xlsx")
    print(f"主报表: {output_file}\n")

    all_rows = parse_all_judge_files(args.input_dir)
    if not all_rows:
        raise SystemExit("没有找到任何评分数据")

    tasks = collect_all_tasks(all_rows)
    if not tasks:
        raise SystemExit("数据中没有任何有效的 task 评分")
    print(f"检测到 task 列表: {tasks}\n")

    print("构建原子桶...")
    atomic, _ = _build_atomic_buckets(all_rows)
    print(f"  ✓ 原子桶构建完成: {sum(len(v) for v in atomic.values())} 个 (task, atomic_key) 组\n")

    print("构建主报表...")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_ov, ov_h1, ov_h2 = build_task_overview_sheet(atomic, tasks, args.save_ap)
        if not df_ov.empty:
            write_summary_sheet(writer, df_ov, ov_h1, ov_h2, sheet_name="任务总览")
            print("  ✓ 任务总览已写入")
        for task_name in tasks:
            df_t, t_h1, t_h2 = build_task_summary_sheet(atomic, task_name, args.save_ap)
            if df_t.empty:
                continue
            sheet_name = safe_sheet_name(f"总分_{task_name}")
            write_summary_sheet(writer, df_t, t_h1, t_h2, sheet_name=sheet_name)
            print(f"  ✓ {sheet_name}")
    print(f"\n✓ 主报表已保存: {output_file}")

    if not args.no_split_by_type:
        by_type_dir = os.path.join(os.path.dirname(output_file), BY_TYPE_DIR_NAME)
        print(f"\n按图表类型拆分到: {by_type_dir}")
        write_by_chart_type_excel(atomic, by_type_dir, tasks, args.save_ap)

    if not args.no_detail:
        detail_dir = os.path.join(os.path.dirname(output_file), DETAIL_DIR_NAME)
        print(f"\n生成详细分类结果到: {detail_dir}")
        write_detail_excel(atomic, detail_dir, tasks, args.save_ap)
        print("✓ 详细分类结果已保存")

    print(f"\n{'=' * 72}")
    print("分析完成!")
    print(f"{'=' * 72}")


if __name__ == "__main__":
    main()
